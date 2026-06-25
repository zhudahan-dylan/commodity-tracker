"""
Storage — Excel read/write with append-only semantics, auto-sheet creation,
summary rebuild, and chart refresh.
"""

import os
from datetime import datetime, date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import (
    OUTPUT_DIR, EXCEL_FILE,
    SUMMARY_HEADERS, DATA_HEADERS,
)
from chart_builder import rebuild_charts

# ── Styles (shared with chart_builder) ─────────────────
HEADER_FONT  = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
DATA_FONT    = Font(name="Microsoft YaHei", size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT_CENTER  = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
ALT_FILL     = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
CAT_FILL     = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
UP_FONT      = Font(name="Microsoft YaHei", size=10, bold=True, color="C00000")
DOWN_FONT    = Font(name="Microsoft YaHei", size=10, bold=True, color="007A33")
FLAT_FONT    = Font(name="Microsoft YaHei", size=10, bold=True, color="7F7F7F")
TITLE_FONT   = Font(name="Microsoft YaHei", size=14, bold=True, color="1F3864")
SECTION_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="1F3864")
GREEN_FILL   = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
RED_FILL     = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
SPARK_CHARS  = "▁▂▃▄▅▆▇"


# ── Helpers ────────────────────────────────────────────

def _ensure_output_dir():
    os.makedirs(OUTPUT_DIR, exist_ok=True)


def _filepath() -> str:
    return os.path.join(OUTPUT_DIR, EXCEL_FILE)


def _style_row(ws, row, num_cols, *, font=None, fill=None, align=None):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        if font:   cell.font = font
        if fill:   cell.fill = fill
        if align:  cell.alignment = align
        cell.border = THIN_BORDER


def _auto_width(ws, num_cols, max_row):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in range(1, min(max_row + 1, 500)):
            val = ws.cell(row=row, column=col).value
            if val:
                length = sum(2 if ord(str(c)) > 127 else 1 for c in str(val))
                max_len = max(max_len, length)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 32)


def _norm_date(val) -> str:
    """Normalize a cell value to YYYY-MM-DD string."""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, date):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()[:10]


# ── Sparkline / trend helpers ──────────────────────────

def _sparkline(prices: list[float]) -> str:
    if not prices or max(prices) == min(prices):
        return "▄" * len(prices)
    mn, mx = min(prices), max(prices)
    rng = mx - mn
    return "".join(SPARK_CHARS[min(int((p - mn) / rng * 7), 6)] for p in prices)


def _trend_arrow(prices: list[float]) -> str:
    if len(prices) < 2:
        return "→"
    first, last = prices[0], prices[-1]
    pct = (last - first) / first * 100 if first else 0
    return "↑" if pct > 0.5 else ("↓" if pct < -0.5 else "→")


def _trend_label(prices: list[float]) -> str:
    if len(prices) < 2:
        return "震荡"
    first, last = prices[0], prices[-1]
    pct = (last - first) / first * 100 if first else 0
    if pct > 1:
        return "上行"
    if pct < -1:
        return "下行"
    return "震荡"


# ═══════════════════════════════════════════════════════
#  Workbook initialisation
# ═══════════════════════════════════════════════════════

def _init_workbook() -> Workbook:
    """Create a fresh workbook with Summary sheet."""
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("汇总", 0)
    _write_summary_headers(ws)
    return wb


def _write_summary_headers(ws):
    """Write the summary sheet header rows (title + column headers)."""
    num_cols = len(SUMMARY_HEADERS)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    ws.cell(row=1, column=1, value="📊 大宗商品价格日报 — 汇总").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = LEFT_CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    ws.cell(row=2, column=1, value="数据来源: 生意社(100ppi.com)").font = Font(name="Microsoft YaHei", size=9, color="666666")
    for c, h in enumerate(SUMMARY_HEADERS, 1):
        ws.cell(row=4, column=c, value=h)
    _style_row(ws, 4, num_cols, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)


def _ensure_category_sheet(wb: Workbook, cat_name: str) -> "Worksheet":
    """Get or create a category sheet with data table headers."""
    if cat_name in wb.sheetnames:
        ws = wb[cat_name]
    else:
        ws = wb.create_sheet(cat_name)
        # Write section title + data headers
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws.cell(row=1, column=1, value=f"📊 {cat_name} — 价格趋势").font = TITLE_FONT
        ws.cell(row=1, column=1).alignment = LEFT_CENTER

        # Leave rows 2-N for chart area, write data headers at a fixed offset
        # We'll place data headers at row 50 to leave room for charts
        data_header_row = 50
        ws.merge_cells(start_row=data_header_row - 1, start_column=1,
                        end_row=data_header_row - 1, end_column=6)
        ws.cell(row=data_header_row - 1, column=1,
                value=f"📋 {cat_name} — 原始数据").font = SECTION_FONT
        for c, h in enumerate(DATA_HEADERS, 1):
            ws.cell(row=data_header_row, column=c, value=h)
        _style_row(ws, data_header_row, len(DATA_HEADERS),
                   font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        ws.freeze_panes = f"A{data_header_row + 1}"
    return ws


def _find_data_bounds(ws):  # -> tuple[int, int] | None
    """
    Find the data table bounds in a category sheet.
    Returns (header_row, last_data_row) or None.
    """
    for row in range(1, ws.max_row + 1):
        a1 = ws.cell(row=row, column=1).value
        if a1 and str(a1).strip() == "日期":
            a2 = ws.cell(row=row, column=2).value
            if a2 and "商品名称" in str(a2):
                header = row
                last = header
                for r in range(header + 1, ws.max_row + 1):
                    if ws.cell(row=r, column=1).value is None:
                        break
                    last = r
                return (header, last)
    return None


# ═══════════════════════════════════════════════════════
#  Append records
# ═══════════════════════════════════════════════════════

def append_records(records: list[dict]) -> int:
    """
    Append records to Excel. Dedup by (date, category, commodity_name).
    Returns number of NEW records added.
    """
    if not records:
        return 0

    _ensure_output_dir()
    fp = _filepath()

    if os.path.exists(fp):
        wb = load_workbook(fp)
    else:
        wb = _init_workbook()

    new_count = 0
    categories_touched: set[str] = set()

    for rec in records:
        cat = rec.get("品类", "未知")
        ws = _ensure_category_sheet(wb, cat)

        # ── Build dedup set from existing data ──
        bounds = _find_data_bounds(ws)
        if bounds:
            _, last_row = bounds
            existing: set[tuple] = set()
            for r in range(bounds[0] + 1, last_row + 1):
                d = _norm_date(ws.cell(row=r, column=1).value)
                n = str(ws.cell(row=r, column=2).value or "")
                existing.add((d, n))
        else:
            existing = set()

        # ── Dedup check ──
        key = (_norm_date(rec.get("日期", "")), str(rec.get("商品名称", "")))
        if key in existing:
            continue

        # ── Append ──
        if bounds:
            next_row = bounds[1] + 1
        else:
            next_row = 51   # first data row after header at row 50

        ws.cell(row=next_row, column=1, value=str(rec.get("日期", "")))
        ws.cell(row=next_row, column=2, value=str(rec.get("商品名称", "")))
        ws.cell(row=next_row, column=3, value=rec.get("价格"))
        ws.cell(row=next_row, column=4, value=str(rec.get("单位", "元/吨")))
        ws.cell(row=next_row, column=5, value=rec.get("七日涨跌幅(%)"))
        ws.cell(row=next_row, column=6, value=str(rec.get("记录时间", "")))
        _style_row(ws, next_row, len(DATA_HEADERS), font=DATA_FONT, align=CENTER,
                   fill=ALT_FILL if next_row % 2 == 0 else None)

        # Update auto-filter
        ws.auto_filter.ref = f"A{50}:F{next_row}"

        existing.add(key)
        new_count += 1
        categories_touched.add(cat)

    # ── Rebuild charts for touched categories ──
    for cat in categories_touched:
        ws = wb[cat]
        try:
            rebuild_charts(ws, cat)
        except Exception as exc:
            print(f"  ⚠️ 重建 {cat} 图表失败: {exc}")

    # ── Rebuild summary sheet ──
    _rebuild_summary(wb)

    wb.save(fp)
    print(f"💾 新增 {new_count} 条记录 (去重后)")
    return new_count


# ═══════════════════════════════════════════════════════
#  Summary rebuild
# ═══════════════════════════════════════════════════════

def _rebuild_summary(wb: Workbook):
    """Rebuild the summary sheet from all category sheets."""
    ws = wb["汇总"]

    # Clear old data (row 5+), unmerge first
    merged_to_unmerge = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row >= 5:
            merged_to_unmerge.append(str(mr))
    for mr_str in merged_to_unmerge:
        ws.unmerge_cells(mr_str)

    for r in range(5, ws.max_row + 1):
        for c in range(1, len(SUMMARY_HEADERS) + 1):
            try:
                ws.cell(row=r, column=c).value = None
            except AttributeError:
                pass  # already unmerged

    # Collect data from each category sheet
    all_records: list[dict] = []
    for sheet_name in wb.sheetnames:
        if sheet_name == "汇总":
            continue
        cat_ws = wb[sheet_name]
        bounds = _find_data_bounds(cat_ws)
        if not bounds:
            continue
        header, last = bounds
        for r in range(header + 1, last + 1):
            d = _norm_date(cat_ws.cell(row=r, column=1).value)
            name = str(cat_ws.cell(row=r, column=2).value or "")
            price = cat_ws.cell(row=r, column=3).value
            unit = str(cat_ws.cell(row=r, column=4).value or "")
            if not d or not name or price is None:
                continue
            try:
                price = float(price)
            except (ValueError, TypeError):
                continue
            all_records.append({
                "日期": d,
                "品类": sheet_name,
                "商品名称": name,
                "价格": price,
                "单位": unit,
            })

    if not all_records:
        return

    # Group by (category, commodity), sort by date
    from collections import defaultdict
    groups: dict[tuple, list[tuple[str, float]]] = defaultdict(list)
    for r in all_records:
        groups[(r["品类"], r["商品名称"], r["单位"])].append((r["日期"], r["价格"]))

    row = 5
    current_cat = None
    num_cols = len(SUMMARY_HEADERS)

    # Sort groups by category then commodity
    sorted_groups = sorted(groups.items(), key=lambda x: (x[0][0], x[0][1]))

    for (cat, name, unit), date_prices in sorted_groups:
        # Category separator row
        if cat != current_cat:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
            ws.cell(row=row, column=1, value=f"▎{cat}").font = Font(
                name="Microsoft YaHei", size=11, bold=True, color="375623")
            ws.cell(row=row, column=1).fill = CAT_FILL
            ws.cell(row=row, column=1).alignment = LEFT_CENTER
            for c in range(1, num_cols + 1):
                ws.cell(row=row, column=c).border = THIN_BORDER
            row += 1
            current_cat = cat

        # Sort by date
        date_prices.sort(key=lambda x: x[0])
        prices = [p for _, p in date_prices]
        dates  = [d for d, _ in date_prices]

        latest = prices[-1] if prices else 0
        prev   = prices[-2] if len(prices) >= 2 else latest
        day_chg = (latest - prev) / prev * 100 if prev else 0

        if day_chg > 0.01:
            chg_text = f"↑ +{day_chg:.2f}%"
        elif day_chg < -0.01:
            chg_text = f"↓ {day_chg:.2f}%"
        else:
            chg_text = "→ 0.00%"

        spark = _sparkline(prices[-7:])   # last 7 days
        trend = _trend_label(prices[-7:])
        arrow = _trend_arrow(prices[-7:])

        ws.cell(row=row, column=1, value="")
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=unit)
        ws.cell(row=row, column=4, value=latest)
        ws.cell(row=row, column=5, value=chg_text)
        ws.cell(row=row, column=6, value=spark)
        ws.cell(row=row, column=7, value=f"{arrow} {trend}")

        _style_row(ws, row, num_cols, font=DATA_FONT, align=CENTER)

        # Color the daily change cell
        chg_cell = ws.cell(row=row, column=5)
        if day_chg > 0.01:
            chg_cell.font = UP_FONT
            chg_cell.fill = GREEN_FILL
        elif day_chg < -0.01:
            chg_cell.font = DOWN_FONT
            chg_cell.fill = RED_FILL
        else:
            chg_cell.font = FLAT_FONT

        trend_cell = ws.cell(row=row, column=7)
        if "上行" in trend:
            trend_cell.font = UP_FONT
        elif "下行" in trend:
            trend_cell.font = DOWN_FONT
        else:
            trend_cell.font = FLAT_FONT

        ws.cell(row=row, column=6).font = Font(name="Microsoft YaHei", size=11)
        row += 1

    # Update date subtitle
    ws.cell(row=2, column=1,
            value=f"更新日期: {datetime.now().strftime('%Y-%m-%d')}  数据来源: 生意社(100ppi.com)")

    _auto_width(ws, num_cols, row - 1)
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 18
    ws.column_dimensions["G"].width = 14


# ═══════════════════════════════════════════════════════
#  Read / stats (public)
# ═══════════════════════════════════════════════════════

def read_all_records(cat_name=None):  # cat_name: str|None -> list[dict]
    """Read all records, optionally filtered by category."""
    fp = _filepath()
    if not os.path.exists(fp):
        return []

    wb = load_workbook(fp)
    records: list[dict] = []
    sheets = [cat_name] if cat_name else [s for s in wb.sheetnames if s != "汇总"]

    for sn in sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        bounds = _find_data_bounds(ws)
        if not bounds:
            continue
        header, last = bounds
        for r in range(header + 1, last + 1):
            rec = {}
            for c, h in enumerate(DATA_HEADERS, 1):
                val = ws.cell(row=r, column=c).value
                if isinstance(val, datetime):
                    val = val.strftime("%Y-%m-%d")
                elif isinstance(val, date):
                    val = val.strftime("%Y-%m-%d")
                rec[h] = val
            rec["品类"] = sn
            records.append(rec)
    wb.close()
    return records


def get_stats() -> dict:
    """Return summary stats."""
    records = read_all_records()
    if not records:
        return {"total_records": 0, "categories": {}, "date_range": None, "file": _filepath()}

    cats: dict[str, int] = {}
    dates: set[str] = set()
    for r in records:
        c = r.get("品类", "未知")
        cats[c] = cats.get(c, 0) + 1
        d = str(r.get("日期", ""))[:10]
        if d:
            dates.add(d)

    sorted_dates = sorted(dates)
    return {
        "total_records": len(records),
        "categories": cats,
        "date_range": (sorted_dates[0], sorted_dates[-1]) if sorted_dates else None,
        "file": _filepath(),
    }
