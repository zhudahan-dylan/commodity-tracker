"""
Chart builder — creates per-unit line charts inside category sheets.

Flow for each category sheet:
  1. Read all records for that category from the data table
  2. Pivot: {date: {commodity_name: price}}
  3. Group commodities by unit
  4. For each unit group, write pivot table + embed line chart
  5. Charts placed above the raw data table
"""

from collections import defaultdict
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import ChartLines
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import CHART_WIDTH, CHART_HEIGHT, SERIES_COLORS

# ── Reusable styles ────────────────────────────────────
HEADER_FONT  = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_FILL  = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
DATA_FONT    = Font(name="Microsoft YaHei", size=10)
CENTER       = Alignment(horizontal="center", vertical="center")
LEFT_CENTER  = Alignment(horizontal="left",   vertical="center")
THIN_BORDER  = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)
ALT_FILL     = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
TITLE_FONT   = Font(name="Microsoft YaHei", size=14, bold=True, color="1F3864")
SECTION_FONT = Font(name="Microsoft YaHei", size=12, bold=True, color="1F3864")


def _style_row(ws, row, num_cols, *, font=None, fill=None, align=None):
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        if font:   cell.font = font
        if fill:   cell.fill = fill
        if align:  cell.alignment = align
        cell.border = THIN_BORDER


def _find_data_table_bounds(ws):  # -> tuple[int, int, int] | None
    """
    Locate the raw data table in a category sheet.
    Returns (header_row, first_data_row, last_data_row) or None.
    We identify the data table by the header row containing '日期'.
    """
    for row in range(1, ws.max_row + 1):
        a1 = ws.cell(row=row, column=1).value
        if a1 and str(a1).strip() == "日期":
            # Make sure column 2 also has a matching header
            a2 = ws.cell(row=row, column=2).value
            if a2 and "商品名称" in str(a2):
                header_row = row
                first_data = row + 1
                # Find last data row
                last_data = first_data
                for r in range(first_data, ws.max_row + 1):
                    if ws.cell(row=r, column=1).value is None:
                        break
                    last_data = r
                return (header_row, first_data, last_data)
    return None


def _read_data_records(ws) -> list[dict]:
    """Read raw records from the data table in a category sheet."""
    bounds = _find_data_table_bounds(ws)
    if not bounds:
        return []
    _, first, last = bounds
    records: list[dict] = []
    for r in range(first, last + 1):
        rec = {
            "日期": ws.cell(row=r, column=1).value,
            "商品名称": ws.cell(row=r, column=2).value,
            "价格": ws.cell(row=r, column=3).value,
            "单位": ws.cell(row=r, column=4).value,
        }
        if rec["日期"] and rec["商品名称"]:
            records.append(rec)
    return records


def _pivot(records):  # -> tuple[list[str], dict, dict]
    """
    Pivot records into chart-friendly structure.
    Returns: (sorted_dates, {commodity_name: {date: price}})
    """
    # Collect all commodities and their units
    comm_units: dict[str, str] = {}
    date_prices: dict[str, dict[str, dict[str, float]]] = defaultdict(
        lambda: defaultdict(dict)
    )  # comm -> date -> price

    all_dates: set[str] = set()
    for r in records:
        d = str(r["日期"])[:10]  # normalise to YYYY-MM-DD
        name = str(r["商品名称"])
        price = r["价格"]
        unit = str(r.get("单位", "元/吨"))
        if price is None:
            continue
        try:
            price = float(price)
        except (ValueError, TypeError):
            continue
        comm_units[name] = unit
        date_prices[name][d] = price
        all_dates.add(d)

    sorted_dates = sorted(all_dates)
    # Build {comm: {date: price}} with None for missing dates
    result = {}  # dict[str, dict[str, float|None]]
    for name in date_prices:
        result[name] = {d: date_prices[name].get(d) for d in sorted_dates}

    return sorted_dates, result, comm_units


def _auto_width(ws, num_cols, max_row):
    for col in range(1, num_cols + 1):
        max_len = 0
        for row in range(1, min(max_row + 1, 400)):
            val = ws.cell(row=row, column=col).value
            if val:
                length = sum(2 if ord(str(c)) > 127 else 1 for c in str(val))
                max_len = max(max_len, length)
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 4, 32)


# ── Public API ─────────────────────────────────────────

def rebuild_charts(ws, cat_name):  # cat_name: str
    """
    Rebuild all charts in a category sheet.
    1. Clear existing chart area (rows 1 to data-table header)
    2. Read data records
    3. Pivot + group by unit
    4. Write pivot tables + embed charts
    """
    bounds = _find_data_table_bounds(ws)
    if not bounds:
        return  # No data table yet

    data_header_row = bounds[0]
    records = _read_data_records(ws)
    if not records:
        return

    sorted_dates, comm_prices, comm_units = _pivot(records)

    # ── Clear old chart area (rows 1 to data_header_row-1) ──
    # Unmerge all merged cells in this range first
    merged_to_unmerge = []
    for mr in ws.merged_cells.ranges:
        if mr.min_row < data_header_row:
            merged_to_unmerge.append(str(mr))
    for mr_str in merged_to_unmerge:
        ws.unmerge_cells(mr_str)

    for r in range(1, data_header_row):
        for c in range(1, 20):
            try:
                ws.cell(row=r, column=c).value = None
            except AttributeError:
                pass

    # ── Remove old charts (prevent accumulation) ──
    ws._charts = []

    current_row = 1

    # ── Title ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1, value=f"📊 {cat_name} — 价格趋势").font = TITLE_FONT
    ws.cell(row=1, column=1).alignment = LEFT_CENTER
    current_row = 3

    # ── One chart per commodity, vertically tiled ──
    all_names = list(comm_prices.keys())
    if not all_names:
        return

    for name in all_names:
        unit = comm_units.get(name, "")

        # --- Sub-title for this commodity ---
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=5)
        sub = ws.cell(row=current_row, column=1,
                      value=f"▎{name} ({unit})")
        sub.font = SECTION_FONT
        sub.alignment = LEFT_CENTER
        current_row += 1

        pivot_start = current_row

        # Header: 日期 | 价格
        ws.cell(row=current_row, column=1, value="日期")
        ws.cell(row=current_row, column=2, value=f"价格({unit})")
        _style_row(ws, current_row, 2, font=HEADER_FONT, fill=HEADER_FILL, align=CENTER)
        current_row += 1

        # Data rows for this commodity
        for di, d in enumerate(sorted_dates):
            val = comm_prices[name].get(d)
            if val is None:
                continue
            ws.cell(row=current_row, column=1, value=d)
            ws.cell(row=current_row, column=2, value=val)
            _style_row(ws, current_row, 2, font=DATA_FONT, align=CENTER,
                       fill=ALT_FILL if di % 2 == 0 else None)
            current_row += 1

        pivot_end = current_row - 1

        # Jump to next row if only 1 data point (avoid empty chart)
        data_rows = pivot_end - pivot_start
        if data_rows < 1:
            current_row += 1
            continue

        # --- Chart for this commodity ---
        chart = LineChart()
        chart.title = name
        chart.width  = 18
        chart.height = 10

        chart.y_axis.delete = False
        chart.x_axis.delete = False
        chart.y_axis.title = unit
        chart.x_axis.title = "日期"
        chart.y_axis.numFmt = '#,##0'
        chart.y_axis.majorTickMark = "out"
        chart.y_axis.minorTickMark = "none"
        chart.y_axis.tickLblPos = "nextTo"
        chart.y_axis.majorGridlines = ChartLines()
        chart.x_axis.majorTickMark = "out"
        chart.x_axis.tickLblPos = "nextTo"
        chart.legend.position = "b"

        cats = Reference(ws, min_col=1, min_row=pivot_start + 1, max_row=pivot_end)
        data_ref = Reference(ws, min_col=2, min_row=pivot_start, max_row=pivot_end)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats)
        if SERIES_COLORS:
            chart.series[0].graphicalProperties.line.solidFill = SERIES_COLORS[0]
        chart.series[0].marker.symbol = "circle"
        chart.series[0].marker.size = 4

        # Place chart to the right of data, anchor at the data start
        chart.anchor = f"D{pivot_start}"
        ws.add_chart(chart, f"D{pivot_start}")

        current_row += 2   # gap between commodities

    # ── Adjust column widths ──
    _auto_width(ws, 5, current_row)
    # Ensure chart columns have room
    for col_letter in ["A", "B", "C"]:
        if ws.column_dimensions[col_letter].width < 14:
            ws.column_dimensions[col_letter].width = 14
